from configparser import ConfigParser
import json
import re
import regex
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import shutil
import os


basedir = os.path.dirname(__file__)
config = ConfigParser()

with open('config.ini', encoding='utf-8') as f:
    config.read_file(f)

kommuneInfo = json.loads(config.get("VARS", "kommuneInfo"))
mappeTyper = json.loads(config.get("VARS", "mappeTyper"))

#For å konvertere nøklene tilbake til int, for må være string i json format
kommuneInfo = {int(k): v for k, v in kommuneInfo.items()}


def find_errors(path, checkKommune, checkDato, checkPersonnr, checkNavn):

    try:
        wb = load_workbook(path)
    except InvalidFileException:
        return "FEIL: Manglende fil eller feil filtype"

    ws = wb.active

    error_string = ""

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):

        kommune = row[0].value
        kommunenr = row[1].value
        archiveCreator = row[2].value
        birthDate = row[3].value
        personnr = row[4].value  # Index 4 corresponds to column E
        lastName = row[5].value
        firstName = row[6].value
        middleName = row[7].value
        noOfDir = row[8].value
        box = row[9].value
        dirType = row[10].value
        morsDate = row[11].value
        dirSharedWith = row[12].value


        if checkKommune:
            error_string += validateKommune(kommune=kommune, kommunenr=kommunenr, row=row[0].row)

        if(archiveCreator is None):
            error_string += f'Manglende arkivskaper på linje {row[0].row}\n'

        if checkDato:
            error_string += validateBirthDate(birthDate=birthDate, row=row[0].row)
        if checkPersonnr:
            error_string += validatePersonnr(personnr=personnr, row=row[0].row)

        if checkNavn:  
            error_string += validate_last_name(lastName, row[0].row)
            error_string += validate_first_name(firstName, row[0].row)
            error_string += validate_middle_name(middleName, row[0].row)
        error_string += validate_no_of_dir(noOfDir, row[0].row)
        error_string += validate_box(box, row[0].row)
        error_string += validate_dir_type(dirType, row[0].row)
        error_string += validate_mors_date(morsDate, row[0].row)
        
    return error_string



def validateKommune(kommune, kommunenr, row):

    valid = True

    text = ""

    if kommune not in kommuneInfo.values():
        text += f'Feil eller manglende kommune på rad: {row}\n'
        valid = False

    if kommunenr not in kommuneInfo.keys():
        text += f'Feil eller manglende kommunenummer på rad: {row}\n'
        valid = False

    if not valid:
        return text

    if (kommune != kommuneInfo.get(kommunenr)):
        text += f'Kommune har feil kommunenummer på rad: {row}\n'

    return text


def validateBirthDate(birthDate, row):

    if(isinstance(birthDate, datetime)):
        birthDate = birthDate.strftime('%d%m%Y')

    if(birthDate is not None):
        birthDate = str(birthDate).strip(" ")

    if birthDate is None or len(birthDate) == 0:
        return f'Manglende fødseldato på rad: {row}\n'
        

    elif validateDateFormat(birthDate):
        return f'Feil format på fødselsdato på rad: {row}\n'
    
    return ""

def validateDateFormat(date):
    return not re.match(r'^(0[1-9]|[12][0-9]|3[01])(0[1-9]|1[0-2])\d{4}$', date)

def validatePersonnr(personnr, row):
    if(personnr is None):
        return f'Manglende personnummer på rad: {row}\n'
    elif not re.match(r'^\d{5}$', str(personnr)):
        return f'Feil personnummer på rad: {row} Verdi: {personnr}\n'
    
    return ""

navnFormat = r'^[\p{L}]+([\-\' ][\p{L}]+)*$'

def validate_last_name(lastName, row):
    text = ""
    if lastName:
        lastName = lastName.strip()
    if lastName is None or lastName == "":
        text += f'Manglende etternavn på rad: {row}\n'
    elif not regex.fullmatch(navnFormat, lastName):
        text += f'Feil format på etternavn på rad: {row}\n'
    return text

def validate_first_name(firstName, row):
    text = ""
    if firstName:
        firstName = firstName.strip()
    if firstName is None or firstName == "":
        text += f'Manglende fornavn på rad: {row}\n'
    elif not regex.fullmatch(navnFormat, firstName):
        text += f'Feil format på forrnavn på rad: {row}\n'
    return text

def validate_middle_name(middleName, row):
    text = ""
    if middleName:
        middleName = middleName.strip()
        if not regex.fullmatch(navnFormat, middleName):
            text += f'Mellomnavn på feil format på rad: {row}\n'
    return text

def validate_no_of_dir(noOfDir, row):
    pattern = r'^[1-9][0-9]*$'
    if not noOfDir:
        return f'Manglende Antall Mapper på rad: {row}\n'
    if not re.fullmatch(pattern, str(noOfDir)):
        return f'Feil format på Antall Mapper på rad: {row}\n'
    
    return ""
    
def validate_box(box, row):
    pattern = r'^[1-9][0-9]*$'
    if not box:
        return f'Manglende boks på rad: {row}\n'
    if not re.fullmatch(pattern, str(box)):
        return f'Feil format på boks på rad: {row}\n'
    
    return ""
    
def validate_dir_type(dirType, row):
    if dirType not in mappeTyper:
        return f'Feil eller manglende mappetype på rad: {row}\n'
    return ""
    
def validate_mors_date(morsDate, row):
    if not morsDate:
        return ""
    if validateDateFormat(morsDate):
        return f'Feil morsdato på rad: {row}\n'



def fix_errors(path):
    try:
        copy_path = copy_excel_file(path)
    except FileNotFoundError:
        return "FEIL: Manglende fil"
    except PermissionError:
        return "FEIL: Korrigert fil allerede åpen.\nVennligst lukk filen og prøv igjen."
    

    try:
        wb = load_workbook(path)
    except InvalidFileException:
        return "FEIL: Manglende fil eller feil filtype"
    ws = wb.active
    text = ""
    output_text = ""

    #Korrigerer dato
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):

        birthDate = row[3].value
        personnr = row[4].value
        lastName = row[5].value
        firstName = row[6].value
        middleName = row[7].value
        morsDate = row[11].value
        

        if(isinstance(birthDate, datetime)):
            birthDate = birthDate.strftime('%d%m%Y')

        if birthDate is not None:
            

            birthDate = str(birthDate)

            if(validateDateFormat(birthDate)):
                fixed, message = fix_date(birthDate, row[0].row)
                ws.cell(row=row[0].row, column=4).value = fixed
                output_text += message

        if(lastName):
            if(validate_last_name):
                fixed_lname, message = clean_name(lastName, row[0].row)
                ws.cell(row=row[0].row, column=6).value = fixed_lname
                output_text += message

        if(firstName):
            if(validate_first_name):
                fixed_fname, message = clean_name(firstName, row[0].row)
                ws.cell(row=row[0].row, column=7).value = fixed_fname
                output_text += message
        if(middleName):
            if(validate_middle_name):
                fixed_mname, message = clean_name(middleName, row[0].row)
                ws.cell(row=row[0].row, column=8).value = fixed_mname
                output_text += message
        
        if(morsDate):
            fixed, message = fix_date(morsDate, row[0].row)
            ws.cell(row=row[0].row, column=12).value = fixed
            output_text += message

    wb.save(copy_path)
    text += "Korrigert fil lagret til: " + copy_path + "\n"
    text += output_text
    return text



def copy_excel_file(source_path):
    """
    AI-generert
    Copies an Excel file and adds '_korrigert' before the file extension.
    The copy is saved in the same directory as the original.

    :param source_path: Path to the original Excel file.
    :return: Path to the copied file.
    """
    if not os.path.isfile(source_path):
        raise FileNotFoundError(f"File not found: {source_path}")

    base, ext = os.path.splitext(source_path)
    copy_path = f"{base}_korrigert{ext}"

    shutil.copy2(source_path, copy_path)
    return copy_path


def fix_date(date, row):

    date = strip_non_numeric(date)

    message = ""
    
    if(date and validateDateFormat(date)):
        message = f'Kunne ikke korrigere: {date} på rad: {row} \n'



    return date, message


def strip_non_numeric(s):
    return re.sub(r'\D', '', s)  # \D matches any non-digit character

def clean_name(s, row):
    # Keep only allowed characters
    s = regex.sub(r"[^\p{L}\-'. ]", '', s)
    # Collapse repeated dashes, spaces, or periods
    s = re.sub(r'[-]{2,}', '-', s)
    s = re.sub(r'[.]{2,}', '.', s)
    s = re.sub(r'[ ]{2,}', ' ', s)
    s = re.sub(r'\s*-\s*', '-', s)

    message = ""

    if(validate_middle_name(s, row)):
        message = f'Kunne ikke korrigere: {s} på rad {row}\n'

    # Trim unwanted characters at ends
    return s.strip(' -'), message


# fix_errors('D:/Viktor/Bodø kommune. Barnehagekontoret. Fa 1-45. AKS_23_157_7.xlsx')
# fix_errors('D:/Viktor/Narvik kommune. Skolestyre.Fa 1-180. AKS_20_137.xlsx')
# fix_errors('D:/Viktor/Vefsn Kommune HR Avdeling Personal sortert stigende på etternavn.xlsx')
# print(find_errors('D:/Viktor/Bodø kommune. Barnehagekontoret. Fa 1-45. AKS_23_157_7.xlsx'))
# find_errors('D:/Viktor/Narvik kommune. Skolestyre.Fa 1-180. AKS_20_137.xlsx')
# find_errors('D:/Viktor/Vefsn Kommune HR Avdeling Personal sortert stigende på etternavn.xlsx')