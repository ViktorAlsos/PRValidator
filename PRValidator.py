import openpyxl
import re
from openpyxl import load_workbook

kommuneInfo = {
    1804: 'Bodø',
    1806: 'Narvik',
    1811: 'Bindal',
    1812: 'Sømna',
    1813: 'Brønnøy',
    1815: 'Vega',
    1816: 'Vevelstad',
    1818: 'Herøy',
    1820: 'Alstahaug',
    1822: 'Leirfjord',
    1824: 'Vefsn',
    1825: 'Grane',
    1826: 'Hattfjelldal',
    1827: 'Dønna',
    1828: 'Nesna',
    1832: 'Hemnes',
    1833: 'Rana',
    1834: 'Lurøy',
    1835: 'Træna',
    1836: 'Rødøy',
    1837: 'Meløy',
    1838: 'Gildeskål',
    1839: 'Beiarn',
    1840: 'Saltdal',
    1841: 'Fauske',
    1845: 'Sørfold',
    1848: 'Steigen',
    1851: 'Lødingen',
    1853: 'Evenes',
    1856: 'Røst',
    1857: 'Værøy',
    1859: 'Flakstad',
    1860: 'Vestvågøy',
    1865: 'Vågan',
    1866: 'Hadsel',
    1867: 'Bø',
    1868: 'Øksnes',
    1870: 'Sortland',
    1871: 'Andøy',
    1874: 'Moskenes',
    1875: 'Hamarøy'
}


def find_errors(path):
    wb = load_workbook(path)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):

        kommune = row[0].value
        kommunenr = row[1].value
        archiveCreator = row[2].value
        birthDate = row[3].value
        personnr = row[4].value  # Index 4 corresponds to column E
        lastName = row[5].value
        firstName = row[6].value
        middleName = row[7].value
        noOfDir = row[8].value
        box = row[9].value
        dirType = row[10].value
        morsDate = row[11].value
        dirSharedWith = row[12].value


        validateKommune(kommune=kommune, kommunenr=kommunenr, rad=row[0].row)



        if(personnr is None):
            print(f'Manglende personnummer på linje {row[0].row}')
        elif not re.match(r'^\d{5}$', str(personnr)):
            print(f'Feil personnummer på linje {row[0].row} Verdi: {personnr}')


def validateKommune(kommune, kommunenr, rad):

    if kommune not in kommuneInfo.values():
        print(f'Feil eller manglende kommune på rad: {rad}')
    if kommunenr not in kommuneInfo.keys():
        print(f'Feil eller manglende kommunenummer på rad: {rad}')

    if (kommune != kommuneInfo.get(kommunenr)):
        print(f'Kommune har feil kommunenummer på rad: {rad}')

    


find_errors('D:/Viktor/Bodø kommune. Barnehagekontoret. Fa 1-45. AKS_23_157_7.xlsx')